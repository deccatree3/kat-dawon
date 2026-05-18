"""청구마감 엑셀 파서.

요약시트에서 문서 메타데이터와 세부 항목을 추출한다.
값과 수식을 동시에 읽어 후속 검수 단계에서 수식 참조 정합성을 검사할 수 있게 한다.
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

import openpyxl


_CELL_REF_RE = re.compile(r"'?([^'!=+\-*/,()\s]+)'?!([A-Z]+)(\d+)")


def _col_letter_to_index(letters: str) -> int:
    n = 0
    for ch in letters:
        n = n * 26 + (ord(ch) - ord("A") + 1)
    return n


# -----------------------------------------------------------------------------
# 데이터 모델
# -----------------------------------------------------------------------------
@dataclass
class ParsedItem:
    row_index: int
    category: Optional[str]
    item_name: Optional[str]
    quantity: Optional[float]
    unit_price: Optional[float]
    amount: Optional[float]
    formula_amount: Optional[str]          # G열 원본 수식
    formula_quantity: Optional[str]        # E열 원본 수식
    remarks: Optional[str]


@dataclass
class ParsedDocument:
    company: str
    year_month: str                 # 'YYYY-MM'
    period_from: Optional[str]
    period_to: Optional[str]
    supply_amount: float
    vat: float
    total_amount: float
    source_file: str
    items: list[ParsedItem] = field(default_factory=list)
    # 검수용 외부 시트 값 캐시 — (sheet_name, row, col) → value
    external_cells: dict = field(default_factory=dict)
    # 시트별 보조 메트릭 — 청구 항목 검증에 사용
    # 예: {'반품택배': {'row_count': 4, 'total_AG': 8000}}
    sheet_metrics: dict = field(default_factory=dict)

    def lookup_value(self, sheet: str, row: int, col: int):
        return self.external_cells.get((sheet, row, col))


# -----------------------------------------------------------------------------
# 파서 상수
# -----------------------------------------------------------------------------
# 업체명 매칭용 — 파일명에 포함되면 해당 업체로 판정
COMPANY_KEYWORDS = {
    "네이처뉴트리션": "네이처뉴트리션",
    "캐처스": "캐처스",
}

# 세부 항목 시작 행 (요약시트 기준)
ITEM_START_ROW = 25

# 컬럼 인덱스 (1-based)
COL_CATEGORY = 2    # B
COL_SUB1 = 3        # C
COL_SUB2 = 4        # D
COL_QTY = 5         # E
COL_UNIT = 6        # F
COL_AMOUNT = 7      # G
COL_REMARK = 8      # H


# -----------------------------------------------------------------------------
# 유틸
# -----------------------------------------------------------------------------
def _detect_company(filename: str) -> str:
    for key, name in COMPANY_KEYWORDS.items():
        if key in filename:
            return name
    return "UNKNOWN"


def _parse_period(text: str) -> tuple[Optional[str], Optional[str], Optional[str]]:
    """'가. 청구 기간 : 2026년01월 01일 ~ 2026년 01월 31일' → (from, to, ym)."""
    if not text:
        return None, None, None
    # 숫자만 추출
    nums = re.findall(r"(\d{4})\s*년\s*(\d{1,2})\s*월\s*(\d{1,2})\s*일", text)
    if len(nums) >= 2:
        y1, m1, d1 = nums[0]
        y2, m2, d2 = nums[1]
        frm = f"{y1}-{int(m1):02d}-{int(d1):02d}"
        to = f"{y2}-{int(m2):02d}-{int(d2):02d}"
        ym = f"{y1}-{int(m1):02d}"
        return frm, to, ym
    return None, None, None


def _cell_v(row: int, col: int):
    return ws.cell(row=row, column=col).value


# -----------------------------------------------------------------------------
# 메인 파서
# -----------------------------------------------------------------------------
def parse_billing_file(path: str | Path) -> ParsedDocument:
    path = Path(path)

    # ── 1단계: 요약시트의 수식만 경량 추출 (data_only=False, read_only) ──
    wb_formulas = openpyxl.load_workbook(path, data_only=False, read_only=True)
    try:
        summary_sheet_name = wb_formulas.sheetnames[0]
        ws_f = wb_formulas[summary_sheet_name]
        formula_qty: dict[int, str] = {}
        formula_amt: dict[int, str] = {}
        for r_idx, row in enumerate(
            ws_f.iter_rows(min_row=1, max_row=80, max_col=COL_AMOUNT, values_only=True),
            start=1,
        ):
            for c_idx, value in enumerate(row, start=1):
                if isinstance(value, str) and value.startswith("="):
                    if c_idx == COL_QTY:
                        formula_qty[r_idx] = value
                    elif c_idx == COL_AMOUNT:
                        formula_amt[r_idx] = value
    finally:
        wb_formulas.close()

    # ── 2단계: 수식이 참조하는 외부 시트 셀 좌표 집합 수집 ──────────────────
    needed_cells: dict[str, set[tuple[int, int]]] = {}
    for formula in list(formula_qty.values()) + list(formula_amt.values()):
        for match in _CELL_REF_RE.finditer(formula):
            sheet, col_letters, row_str = match.group(1), match.group(2), match.group(3)
            col_idx = _col_letter_to_index(col_letters)
            needed_cells.setdefault(sheet, set()).add((int(row_str), col_idx))

    # ── 3단계: 값 모드 read_only 로드 — 요약시트 전체 + 필요한 외부 셀만 추출 ──
    wb_values = openpyxl.load_workbook(path, data_only=True, read_only=True)
    external_cells: dict[tuple[str, int, int], object] = {}
    summary_grid: dict[tuple[int, int], object] = {}
    shipping_count: Optional[int] = None  # 출고택배(cj) 시트의 데이터 행수 = BTOC택배비 출고 건수
    sheet_metrics: dict[str, dict[str, float]] = {}
    try:
        # 요약시트: 1~80행 × 1~8열 전체를 dict로
        ws = wb_values[summary_sheet_name]
        for r_idx, row in enumerate(
            ws.iter_rows(min_row=1, max_row=80, max_col=COL_REMARK, values_only=True),
            start=1,
        ):
            for c_idx, value in enumerate(row, start=1):
                if value is not None:
                    summary_grid[(r_idx, c_idx)] = value

        # 외부 시트: 수식이 참조하는 셀만 read_only iter로 추출
        for sheet_name, coords in needed_cells.items():
            if sheet_name == summary_sheet_name or sheet_name not in wb_values.sheetnames:
                continue
            ws_ext = wb_values[sheet_name]
            max_row_needed = max(r for r, _ in coords)
            max_col_needed = max(c for _, c in coords)
            for r_idx, row in enumerate(
                ws_ext.iter_rows(
                    min_row=1, max_row=max_row_needed, max_col=max_col_needed,
                    values_only=True,
                ),
                start=1,
            ):
                for c_idx, value in enumerate(row, start=1):
                    if (r_idx, c_idx) in coords and value is not None:
                        external_cells[(sheet_name, r_idx, c_idx)] = value

        # 시트별 메트릭 추출 — 시트마다 헤더 위치·키 컬럼이 달라서 매핑 정의
        # (시트 키워드, 데이터 시작 행, 키 컬럼 1-based, 합계 컬럼 1-based or None, 행 필터 함수 or None)
        # 키 컬럼은 '진짜 데이터 행에서만 채워지는' 컬럼을 골라야 정확.
        def _btoc_bundle_filter(v):
            s = str(v) if v is not None else ""
            return ("선물세트" in s) and ("스키니퓨리티" not in s)

        SHEETS_TO_SCAN = [
            ("출고택배",       2, 1, 33,   None),
            ("반품택배",       2, 1, 33,   None),
            ("착불",           2, 2, 3,    None),
            ("부자재 사용내역", 2, 5, None, None),
            ("해외포장비",     2, 2, None, None),
            ("BTOB출고",       2, 3, None, None),
            ("용차",           4, 1, None, None),
            ("임가공",         3, 4, None, None),
            ("입고",           2, 2, None, None),
            ("택배입고",       2, 2, None, None),
            ("클레임",         3, 2, None, None),
            # btoc번들작업: 상품명에 '선물세트' 포함 AND '스키니퓨리티' 미포함 행만 카운트
            ("btoc번들작업",   2, 3, 5,    _btoc_bundle_filter),
            ("btob번들작업",   2, 3, 5,    None),
        ]

        # footer/total 텍스트 — 키 값이 이런 텍스트면 데이터 행으로 인정 X
        _FOOTER_KW = ("합계", "총계", "소계", "total", "TOTAL")

        def _scan_sheet(ws, start_row: int, key_col: int, total_col, key_filter=None):
            cnt = 0
            total = 0.0
            max_col = max(key_col, total_col or 0)
            for row in ws.iter_rows(
                min_row=start_row, max_row=20000, max_col=max_col, values_only=True,
            ):
                v_key = row[key_col - 1] if len(row) >= key_col else None
                # 빈 셀
                if v_key is None:
                    if cnt > 0:
                        break
                    continue
                # 텍스트면 footer 키워드 검사 + 빈 문자열 제외
                if isinstance(v_key, str):
                    s = v_key.strip()
                    if not s or any(kw in s for kw in _FOOTER_KW):
                        if cnt > 0:
                            break
                        continue
                # 추가 필터 (예: 특정 상품명만 카운트)
                if key_filter is not None and not key_filter(v_key):
                    continue
                cnt += 1
                if total_col and len(row) >= total_col:
                    v = row[total_col - 1]
                    if isinstance(v, (int, float)):
                        total += v
            return cnt, total

        # 같은 시트가 여러 키워드에 매칭되는 것 방지 — 첫 매칭 키워드만 사용
        for sn in wb_values.sheetnames:
            if sn == summary_sheet_name:
                continue
            for kw, start_row, key_col, total_col, key_filter in SHEETS_TO_SCAN:
                if kw not in sn:
                    continue
                cnt, total_val = _scan_sheet(
                    wb_values[sn], start_row, key_col, total_col, key_filter
                )
                if cnt > 0:
                    m = {"row_count": float(cnt)}
                    if total_col and total_val > 0:
                        m["total"] = total_val
                    sheet_metrics[sn] = m
                if "출고택배" in kw:
                    shipping_count = cnt if cnt > 0 else None
                # btoc번들작업 시트: 정석 위반 행(비-스키니 & 비-선물세트의 입고수량 행) 카운트.
                # 필터로 거르는 동시에 알람 발생용 메트릭으로 저장.
                if kw == "btoc번들작업":
                    anomaly_cnt = 0
                    anomaly_qty = 0.0
                    samples: list[str] = []
                    for r_idx, row in enumerate(
                        wb_values[sn].iter_rows(
                            min_row=start_row, max_row=200, max_col=5, values_only=True
                        ),
                        start_row,
                    ):
                        name = row[2] if len(row) >= 3 else None
                        in_qty = row[4] if len(row) >= 5 else None
                        if not isinstance(in_qty, (int, float)) or in_qty == 0:
                            continue
                        if not name:
                            continue
                        s = str(name)
                        if "선물세트" in s or "스키니퓨리티" in s:
                            continue
                        anomaly_cnt += 1
                        anomaly_qty += in_qty
                        if len(samples) < 3:
                            samples.append(f"R{r_idx} '{s}' (입고 {in_qty:.0f})")
                    if anomaly_cnt > 0:
                        m = sheet_metrics.setdefault(sn, {})
                        m["anomaly_count"] = float(anomaly_cnt)
                        m["anomaly_qty"] = anomaly_qty
                        # 샘플은 별도 sheet_metric로는 못 저장하므로 파서 로그
                        print(f"[NOTICE] btoc번들작업 정석 위반 {anomaly_cnt}행: {'; '.join(samples)}")
                break
    finally:
        wb_values.close()

    def _cell_v(row: int, col: int):
        return summary_grid.get((row, col))

    company = _detect_company(path.name)

    # ── 청구 기간 / year_month ────────────────────────────────────────────
    period_text = ""
    for r in range(16, 22):
        val = _cell_v(r, 2)
        if val and "청구 기간" in str(val):
            period_text = str(val)
            break
    period_from, period_to, year_month = _parse_period(period_text)

    if not year_month:
        # 파일명에서 fallback: '2026년 01월'
        m = re.search(r"(\d{4})\D+(\d{1,2})", path.name)
        if m:
            year_month = f"{m.group(1)}-{int(m.group(2)):02d}"
        else:
            year_month = "UNKNOWN"

    # ── 합계금액 행 탐색 ────────────────────────────────────────────────
    sum_row = None
    for r in range(ITEM_START_ROW, 80):
        for c in range(1, 10):
            v = _cell_v(r, c)
            if v and "합계금액" in str(v):
                sum_row = r
                break
        if sum_row:
            break

    if sum_row is None:
        raise ValueError(f"'합계금액' 행을 찾을 수 없음: {path.name}")

    # 공급가 / VAT / 청구총액 — 요약시트 기준
    # 합계행(sum_row)에 공급가, sum_row+1에 VAT, sum_row+2에 총액 (G열)
    supply = _cell_v(sum_row, COL_AMOUNT) or 0.0
    vat = _cell_v(sum_row + 1, COL_AMOUNT) or 0.0
    total = _cell_v(sum_row + 2, COL_AMOUNT) or 0.0

    # 항목명 → 세부시트 키워드 매핑 — 요약시트 수량이 없으면 시트 행수로 채움
    ITEM_TO_SHEET_KW = {
        "BTOC택배비": "출고택배",
        "반품 택배": "반품택배",
        "반품택배": "반품택배",
        "임가공": "임가공",
        "용차": "용차",
        "클레임비용": "클레임",
        "택배입고": "택배입고",
        "택배착불": "착불",
    }

    def _qty_from_sheet(item_nm: str) -> Optional[float]:
        kw = ITEM_TO_SHEET_KW.get(item_nm)
        if not kw:
            return None
        for sn, m in sheet_metrics.items():
            if kw in sn:
                rc = m.get("row_count")
                return float(rc) if rc else None
        return None

    # ── 세부 항목 파싱 ───────────────────────────────────────────────────
    items: list[ParsedItem] = []
    current_category: Optional[str] = None

    for r in range(ITEM_START_ROW, sum_row):
        cat = _cell_v(r, COL_CATEGORY)
        if cat and str(cat).strip():
            current_category = str(cat).strip()

        amount = _cell_v(r, COL_AMOUNT)
        qty = _cell_v(r, COL_QTY)
        unit = _cell_v(r, COL_UNIT)

        # 금액이 숫자이고 0 이상인 경우에만 항목으로 기록
        if not isinstance(amount, (int, float)):
            continue

        # 항목명 조립: D열(세부항목)이 대표, 없으면 C열, 그 다음 B열
        sub1 = _cell_v(r, COL_SUB1)
        sub2 = _cell_v(r, COL_SUB2)
        name_parts = [str(v) for v in (sub2, sub1) if v is not None and str(v).strip()]
        item_name = name_parts[0] if name_parts else (str(cat) if cat else "")

        remark = _cell_v(r, COL_REMARK)

        # 요약시트에 수량이 비어 있는 항목은 세부시트 행수로 보완
        final_qty = float(qty) if isinstance(qty, (int, float)) else None
        if final_qty is None and item_name:
            inferred = _qty_from_sheet(item_name)
            if inferred:
                final_qty = inferred

        items.append(
            ParsedItem(
                row_index=r,
                category=current_category,
                item_name=str(item_name) if item_name else None,
                quantity=final_qty,
                unit_price=float(unit) if isinstance(unit, (int, float)) else None,
                amount=float(amount),
                formula_amount=formula_amt.get(r),
                formula_quantity=formula_qty.get(r),
                remarks=str(remark) if remark else None,
            )
        )

    return ParsedDocument(
        company=company,
        year_month=year_month,
        period_from=period_from,
        period_to=period_to,
        supply_amount=float(supply),
        vat=float(vat),
        total_amount=float(total),
        source_file=path.name,
        items=items,
        external_cells=external_cells,
        sheet_metrics=sheet_metrics,
    )


# -----------------------------------------------------------------------------
# 보관비 시트 파서
# -----------------------------------------------------------------------------
def parse_storage_sheet(path: str | Path) -> list[dict]:
    """보관비 시트에서 상품별 재고/PLT 데이터를 추출한다.

    PLT 컬럼(G)의 merged cell 그룹을 파악해 각 행에 plt_group 식별자를 부여한다.
    예: G40:G49가 merge되어 PLT=2면, R40~R49는 같은 그룹("G40-G49")에 속하고
        그룹 전체가 2 PLT를 공유한다는 의미.

    Returns:
        list of dict with keys:
            product_code, product_name, damage_flag, loc_group, location,
            quantity, pallet_count, plt_group
    """
    path = Path(path)
    # merged_cells 정보를 위해 read_only=False로 열기 (데이터양 적어 부하 작음)
    wb = openpyxl.load_workbook(path, data_only=True)
    try:
        sheet_name = None
        for sn in wb.sheetnames:
            if "보관" in sn:
                sheet_name = sn
                break
        if sheet_name is None:
            return []

        ws = wb[sheet_name]
        # PLT 컬럼(G=7)의 merged 범위만 미리 dict로 (row → "G{min}-G{max}")
        plt_merged: dict[int, str] = {}
        for mr in ws.merged_cells.ranges:
            if mr.min_col == 7 and mr.max_col == 7:
                gid = f"G{mr.min_row}-G{mr.max_row}"
                for r in range(mr.min_row, mr.max_row + 1):
                    plt_merged[r] = gid

        rows: list[dict] = []
        for r_idx in range(3, 2000):
            code = ws.cell(row=r_idx, column=1).value
            if code is None:
                # 데이터 끝 판단: 다음 5행도 모두 비면 종료
                if all(ws.cell(row=r_idx + i, column=1).value is None for i in range(5)):
                    break
                continue
            name = ws.cell(row=r_idx, column=2).value
            flag = ws.cell(row=r_idx, column=3).value
            loc_grp = ws.cell(row=r_idx, column=4).value
            loc = ws.cell(row=r_idx, column=5).value
            qty = ws.cell(row=r_idx, column=6).value
            plt = ws.cell(row=r_idx, column=7).value
            # 단일 셀 PLT면 그 자체가 그룹 (그룹 ID = "G{r}-G{r}")
            grp = plt_merged.get(r_idx, f"G{r_idx}-G{r_idx}")
            rows.append(
                {
                    "product_code": str(code),
                    "product_name": str(name)[:80] if name else None,
                    "damage_flag": str(flag) if flag else "정상품",
                    "loc_group": str(loc_grp) if loc_grp else None,
                    "location": str(loc) if loc else None,
                    "quantity": float(qty) if isinstance(qty, (int, float)) else 0,
                    "pallet_count": float(plt) if isinstance(plt, (int, float)) else 0,
                    "plt_group": grp,
                }
            )
        return rows
    finally:
        wb.close()
