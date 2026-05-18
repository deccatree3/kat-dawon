"""네이처뉴트리션 BTOB 출고비 검증 — 확장주문검색·박스입수량 기반 정상 박스수 산출.

처리 흐름:
1. 검토 파일(네뉴 물류비 검토-BTOB출고.xlsx)에서 박스 입수량 마스터 추출 → DB 저장
2. 매월 확장주문검색(.xls) 파싱 → [밀크런] 제외 → 상품별 입수량으로 박스수 산출 → 합산
3. 청구마감 요약시트의 BTOB 박스 청구 수량과 비교 (audit_sheets에서 호출)
"""
from __future__ import annotations

import sqlite3
from pathlib import Path
from typing import Optional

# 확장주문검색 파일 컬럼 (1-based)
COL_PRODUCT = 1     # A 상품명
COL_ORDER_QTY = 2   # B 주문수량
COL_RECEIVER = 8    # H 수령자이름 (밀크런 식별용)


def extract_box_master_from_review_file(path: str | Path) -> list[dict]:
    """검토 파일의 Q-R열(컬럼 17, 18)에서 상품별 입수량 마스터 추출."""
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        items: list[dict] = []
        for row in ws.iter_rows(min_row=2, max_row=2000, max_col=18, values_only=True):
            name = row[16] if len(row) >= 17 else None  # Q열 (idx 16)
            cap = row[17] if len(row) >= 18 else None   # R열 (idx 17)
            if not name or not isinstance(cap, (int, float)):
                continue
            items.append({"product_name": str(name).strip(), "units_per_box": float(cap)})
        return items
    finally:
        wb.close()


def parse_extended_orders(path: str | Path) -> list[dict]:
    """확장주문검색 .xls 파일에서 주문 행 파싱.
    Returns: [{'product_name', 'order_qty', 'is_milkrun', 'year_month'}]
    """
    import xlrd
    wb = xlrd.open_workbook(str(path))
    sh = wb.sheet_by_index(0)
    rows: list[dict] = []
    if sh.nrows < 2:
        return rows
    # 첫 행이 헤더라고 가정 (R1: '상품명', '주문수량' 등)
    for r in range(1, sh.nrows):
        try:
            name = sh.cell_value(r, COL_PRODUCT - 1)
            qty = sh.cell_value(r, COL_ORDER_QTY - 1)
            receiver = sh.cell_value(r, COL_RECEIVER - 1)
            order_date = sh.cell_value(r, 3)  # 발주일 D열
        except IndexError:
            continue
        if not name or not isinstance(qty, (int, float)) or qty <= 0:
            continue
        is_milk = bool(receiver and "[밀크런]" in str(receiver))
        # 발주일 → 년월 추출
        ym = None
        if isinstance(order_date, str) and len(order_date) >= 7:
            ym = order_date[:7]  # 'YYYY-MM'
        rows.append({
            "product_name": str(name).strip(),
            "order_qty": float(qty),
            "is_milkrun": is_milk,
            "year_month": ym,
        })
    return rows


def parse_btob_bundle_input(path: str | Path) -> tuple[float, dict]:
    """쿠팡 재고이동건 .xlsx — 모든 시트(form 제외)의 입고수량(D열) 합계.

    Returns: (total_input, per_sheet)
    """
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        total = 0.0
        per_sheet: dict[str, float] = {}
        for sn in wb.sheetnames:
            if sn.lower() == "form":
                continue
            ws = wb[sn]
            s = 0.0
            for row in ws.iter_rows(min_row=2, max_row=500, max_col=4, values_only=True):
                v = row[3] if len(row) >= 4 else None
                if isinstance(v, (int, float)):
                    s += v
            per_sheet[sn] = s
            total += s
        return total, per_sheet
    finally:
        wb.close()


def parse_btoc_bundle_orders(path: str | Path) -> float:
    """확장주문검색 번들작업 검수용 .xls — M열(주문수량) 합계.

    제외 룰: 판매처 상품명(J열, idx 9)에 '스키니퓨리티' 포함 행 제외
    (스키니퓨리티 라인은 BTOB 번들작업비에 미포함 처리)
    """
    import xlrd
    wb = xlrd.open_workbook(str(path))
    sh = wb.sheet_by_index(0)
    total = 0.0
    for r in range(1, sh.nrows):
        try:
            sale_name = sh.cell_value(r, 9)
            qty = sh.cell_value(r, 12)
        except IndexError:
            continue
        if sale_name and "스키니퓨리티" in str(sale_name):
            continue
        if isinstance(qty, (int, float)) and qty > 0:
            total += qty
    return total


def calculate_settlement_box_count(
    orders: list[dict], capacity: dict[str, float], year_month: Optional[str] = None
) -> tuple[float, dict]:
    """정산 대상(밀크런 제외) 박스수 합계 + 상세 정보.

    박스수 = 주문수량 / 입수량. 입수량 0/None 인 항목은 0박스 처리.
    Returns: (total_boxes, details)
      details: {
        'total_orders': N,
        'milkrun_excluded': N,
        'settled_orders': N,
        'missing_capacity': [상품명...],
      }
    """
    total_boxes = 0.0
    missing: set[str] = set()
    cnt_total = cnt_milk = cnt_settled = 0
    for o in orders:
        if year_month and o.get("year_month") and o["year_month"] != year_month:
            continue
        cnt_total += 1
        if o["is_milkrun"]:
            cnt_milk += 1
            continue
        cnt_settled += 1
        cap = capacity.get(o["product_name"])
        if not cap or cap <= 0:
            missing.add(o["product_name"])
            continue
        total_boxes += o["order_qty"] / cap
    return total_boxes, {
        "total_orders": cnt_total,
        "milkrun_excluded": cnt_milk,
        "settled_orders": cnt_settled,
        "missing_capacity": sorted(missing),
    }
